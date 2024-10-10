import openpyxl 

Excel_Leer_Escribir = "C:\\Users\\Public\\oet\\GPS\\operadores\\prueba.xlsx"
sheet = "Unidades"
wb = openpyxl.load_workbook(Excel_Leer_Escribir)
ws = wb[sheet]
row_num = ws.max_row
col_num = ws.max_column
print(row_num, "y", col_num)
row = 2
print("Cabezal = ", ws.cell(row, 1).value)
print("GPS = ", ws.cell(row, 2).value)
print("URL = ", ws.cell(row, 3).value)
print("Cliente = ", ws.cell(row, 4).value)
print("usuario = ", ws.cell(row, 5).value)
print("Contrasena = ", ws.cell(row, 6).value)
print("Integrado = ", ws.cell(row, 7).value)


        
